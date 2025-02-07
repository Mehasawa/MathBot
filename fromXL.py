import openpyxl
from openpyxl import utils
from PIL import Image as pil_image  # Чтобы не было конфликта имен
import os

output_dir = "C:\\Users\\User\\Downloads\\OUTPUT"  # Папка для сохранения картинок
excel_file = 'C:\\Users\\User\\Downloads\\Telegram Desktop\\Новая таблица.xlsx'
workbook = openpyxl.load_workbook(excel_file, data_only=True)

def extract_data_and_images(excel_file, output_dir):

    # Создаем папку для сохранения картинок, если ее нет
    if not os.path.exists(output_dir):
        os.makedirs(output_dir)

    # Загружаем Excel файл
    workbook = openpyxl.load_workbook(excel_file)

    # Перебираем листы (sheets)
    for sheet_name in workbook.sheetnames:
        sheet = workbook[sheet_name]

        # Перебираем рисунки на листе
        for drawing in sheet._images:
            # Получаем координаты ячейки, где находится картинка
            cell = drawing.anchor._from

            # Формируем имя файла для сохранения картинки
            image_name = f"{sheet.title}_{cell.col+1}{cell.row+1}.png"
            image_path = os.path.join(output_dir, image_name)

            # Получаем саму картинку
            image = pil_image.open(drawing.ref)
            # Сохраняем картинку как PNG
            image.save(image_path, "PNG")

            print(f"Изображение сохранено: {image_path}")

        # Вывод текстовых данных из ячеек
        for row in sheet.iter_rows():
            row_values = [cell.value for cell in row]
            print(row_values)


extract_data_and_images(excel_file, output_dir)

